# -*- coding: utf-8 -*-
import json
import re
import sys
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

import openpyxl


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def parse_year(text: str) -> str:
    match = re.search(r"（(\d{4})）", text or "")
    return match.group(1) if match else ""


def split_genres(value: str):
    return [item.strip() for item in re.split(r"/|｜|\|", value or "") if item and item.strip()]


def load_movies(movie_path: Path):
    wb = openpyxl.load_workbook(movie_path, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [clean(item) for item in next(rows)]
    idx = {name: i for i, name in enumerate(headers)}

    movies = []
    for row in rows:
        title = clean(row[idx["电影名称"]])
        if not title:
            continue

        clickobj = clean(row[idx["clickobj"]])
        director = clean(row[idx["导演"]])
        english_title = clean(row[idx["电影英文名称"]])
        actor1 = clean(row[idx["演员1"]])
        actor2 = clean(row[idx["演员2"]])
        genres = split_genres(clean(row[idx["类型"]]))
        score = clean(row[idx["filmscore"]])
        url = clean(row[idx["video_btn"]])

        movies.append(
            {
                "title": title,
                "english_title": english_title,
                "director": director,
                "actors": [item for item in [actor1, actor2] if item],
                "genres": genres,
                "score": score,
                "year": parse_year(clickobj),
                "source_url": url,
            }
        )
    return movies


def load_edges(edge_path: Path):
    wb = openpyxl.load_workbook(edge_path, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [clean(item) for item in next(rows)]
    idx = {name: i for i, name in enumerate(headers)}

    edges = []
    for row in rows:
        source = clean(row[idx["Source"]])
        target = clean(row[idx["Target"]])
        year = clean(row[idx["film_year"]])
        if not source or not target:
            continue
        edges.append({"source": source, "target": target, "year": year})
    return edges


def main():
    base = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("data")
    movie_file = None
    for file in base.glob("*.xlsx"):
        if "net_edges" not in file.name:
            movie_file = file
            break

    edge_file = base / "net_edges.xlsx"
    if movie_file is None or not edge_file.exists():
        raise FileNotFoundError("Dataset files not found in data directory.")

    payload = {
        "movies": load_movies(movie_file),
        "edges": load_edges(edge_file),
    }
    sys.stdout.buffer.write(json.dumps(payload, ensure_ascii=False).encode("utf-8"))


if __name__ == "__main__":
    main()
